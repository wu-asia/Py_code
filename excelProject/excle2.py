from openpyxl import Workbook

wb = Workbook()
ws = wb.active

#ws['a6'] = 'hello'
c1 = ws.cell(1, 6, 'hello')
c2 = ws.cell(1, 7, 'world')
print(c1.value, c2.value, sep=' ', end='\n')
print(c2.value, c2.coordinate, c2.row, c2.column, c2.column_letter)
print(ws.title)
x = 0
for i in range(1, 11):
    for j in range(1, 6):
        ws.cell(i, j, x)
        x += 1


# print(ws['a:c'])
# print(ws[1:3])

for cells in ws.iter_rows():
    for cell in cells:
        print(cell.value, end=' ')
    print('\n')
ws.move_range('4:4', 2, 2)
wb.save('hello.xlsx')

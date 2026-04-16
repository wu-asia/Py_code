from openpyxl import Workbook
from openpyxl.utils import FORMULAE

wb = Workbook()
ws = wb.active

# print(len(FORMULAE))
# print('SUM' in FORMULAE)

ws.append(['价格1', '价格2', '求和', '平均值'])
ws.append([54, 21])
ws.append([12, 64])
ws.append([62, 14])

ws['c2'] = "=SUM(A2:b2)"
ws['d2'] = '=AVERAGE(A2:B2)'

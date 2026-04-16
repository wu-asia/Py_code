from openpyxl import Workbook

wb = Workbook()

ws1 = wb.active

print(ws1.title)

#print(type(ws1))

print(wb.sheetnames)




ws2 = wb.create_sheet('sheet2', 1)
ws3 = wb.create_sheet('sheet3', 2)
# ws4 = wb['sheet3']

# print(ws4.title)
print(wb.sheetnames)
wb.move_sheet('sheet3', -1)

print(wb.sheetnames)

wb.move_sheet('Sheet', 2)

print(wb.sheetnames)

wb['Sheet'].title = 'sheet1'

print(wb.sheetnames)

